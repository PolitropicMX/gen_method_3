from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# PRODUCTOS

# Pictogramas
# Atencion: a
# Daño ambiental: m
# inflamabilidad: i
# Corrosividad: c
# Toxicidad: t

# DICCIONARIO
# A = {KEY: VALUE}
# ACCEDER AL "KEY" Y AL "VALUE"
# A = {KEY: {KEY: VALUE}}

# CONSIDERACIONES AL AGREGAR REACTIVOS
# SOLO 3 PICTOGRAMAS POR REACTIVO

class Productos:
    def __init__(self):
        self.productos = {
##        "Prod 1":{
##            "15-":{
##                "descripcion": "Liquido incoloro; viscosidad Baja; libre de particulas",
##                "peligro":"0",
##                "indicacion":"Pese el 15-340-1 y con polipasto/plataforma añadalo a la tina/tambo",
##                "revision":"verifique que el liquido este libre de particulas"
##                },
##            "Paso 1":{
##                "texto": "texto"
##                },
##            "15-":{
##                "descripcion": "Liquido incoloro; viscosidad Baja; libre de particulas",
##                "peligro":"0",
##                "indicacion":"Pese el 15-340-1 y con polipasto/plataforma añadalo a la tina/tambo",
##                "revision":"verifique que el liquido este libre de particulas"
##                },
##            "Paso 2":{
##                "texto": "texto"
##                }
##            }
##        
        "HB12": { # PRODUCTO 1 # Clave 21-143
            "banner 1":"b2",
            "banner 2":"b8",
            "Paso 1":{
                "texto": "Paso 1) En una tina limpia y adecuada a la carga, agregar los siguientes reactivos" 
                },
            "15-340-1":{
                "descripcion": "Liquido incoloro; viscosidad Baja; libre de particulas",
                "peligro":"m",
                "indicacion":"Pese el 15-340-1 y con polipasto/plataforma añadalo a la tina/tambo",
                "revision":"verifique que el liquido este libre de particulas"
                },
            "15-342":{
                "descripcion": "Liquido incoloro; viscosidad Baja; libre de particulas",
                "peligro":"c",
                "indicacion":"Pese el 15-342 y con polipasto/plataforma añadalo a la tina/tambo",
                "revision":"verifique que el liquido este libre de particulas"
                },
            "15-669":{
                "descripcion": "Liquido incoloro; viscosidad Baja; libre de particulas",
                "peligro":"m",
                "indicacion":"Pese el 15-340-1 y con polipasto/plataforma añadalo a la tina/tambo",
                "revision":"verifique que el liquido este libre de particulas"
                },
            "Paso 2":{
                "texto":"Paso 2) Con un patin llevar la tina con 15-342 y 15-340-1 al area de dispersores, y traer cubeta con 15-699 ya pesado. "
                },
            "Paso 3":{
                "texto":"Paso 3) Procure limpiar con thinner la tuerca, rondanas y propela antes de posicionar la tina."
                },
            "Paso 4":{
                "texto":"Paso 4) Una vez limpia y ensamblada la propela, posicionar la tina debajo de esta.                                                                                          3.- Utilizar una plataforma para vaciar los siguientes reactivos y tener bascula y recipientes listo para hacer las adiciones:"
                },
            "15-323":{
                "descripcion": "Liquido incoloro; viscosidad Baja; libre de particulas",
                "peligro":"am",
                "indicacion":"Pese el 15-323 y con polipasto/plataforma añadalo a la tina/tambo",
                "revision":"verifique que el liquido este libre de particulas"
                },
            "15-665":{
                "descripcion": "Liquido incoloro; viscosidad Baja; libre de particulas",
                "peligro":"0",
                "indicacion":"Pese el 15-665 y con polipasto/plataforma añadalo a la tina/tambo",
                "revision":"verifique que el liquido este libre de particulas"
                },
            "Paso 5":{
                "texto":"Paso 5)Preparar una cubeta o un recipiente grande para mezclar por separado los siguientes reactivos de forma manual. Una vez bien incorporados agregar a la tina"
                },
            "15-627":{
                "descripcion": "Liquido incoloro; viscosidad Baja; libre de particulas",
                "peligro":"ia",
                "indicacion":"Pese el 15-627 y con polipasto/plataforma añadalo a la tina/tambo",
                "revision":"verifique que el liquido este libre de particulas"
                },
            "15-610":{
                "descripcion": "Liquido incoloro; viscosidad Baja; libre de particulas",
                "peligro":"ia",
                "indicacion":"Pese el 15-610 y con polipasto/plataforma añadalo a la tina/tambo",
                "revision":"verifique que el liquido este libre de particulas"
                },
            "Equipo a utilizar": {
                "Equipo a utilizar":"Tina de Acero Inox, Dispersor o Tambo de proceso" 
                },
            "Envase":{
                "Envase_producto":"TAMBOR REUSO AZUL 15-101 P 15-205/CUBETA DE PLASTICO LITOGRAFIADA" ,
                "Envase_control":"Porron de plastico"
                }
            }, # Producto 2
        "HUV-128-B":{ # 22-313-1
            "banner 1":"b2",
            "banner 2":"b8",
            "Paso 1":{
                "texto":"Paso 1) En una tina/tambo/reactor  en area de reactores agregar los siguientes reactivos"
                },
            "15-340-1":{
                "descripcion": "Liquido incoloro; viscosidad Baja; libre de particulas",
                "peligro":"m",
                "indicacion":"Cuando se fabrique en dispersor iniciar agitacion a bajas r.p.m.",
                "revision":"En temporada de frio, se recomienta precalentar a 50-60°C. No Calentar mucho, riesgo de producto oxidado. "
                },
            "15-342":{
                "descripcion": "Liquido incoloro; viscosidad Baja; libre de particulas",
                "peligro":"c",
                "indicacion":"",
                "revision":"Revise que el material es claro, de lo contrario avise al SPR"
                },
            "15-669":{
                "descripcion": "Liquido incoloro; viscosidad Baja; libre de particulas",
                "peligro":"m",
                "indicacion":"",
                "revision":"Revise que el material es claro, de lo contrario avise al SPR"
                },
            "15-323":{
                "descripcion": "Liquido incoloro; viscosidad Baja; libre de particulas",
                "peligro":"am",
                "indicacion":"Una vez terminada la adición, mezcle durante 4-6 minutos",
                "revision":"Revise que el material es claro, de lo contrario avise al SPR"
                },
            "Paso 2":{
                "texto":"Paso 2) Preparar una cubeta/porron para mezclar por separado los siguientes reactivos de forma manual. Una vez bien incorporados agregar a la tina/tambo/reactor"
                },
            "15-627":{
                "descripcion": "Liquido incoloro; viscosidad Baja; libre de particulas",
                "peligro":"ia",
                "indicacion":"",
                "revision":""
                },
            "15-610":{
                "descripcion": "Liquido incoloro; viscosidad Baja; libre de particulas",
                "peligro":"ia",
                "indicacion":"",
                "revision":""
                },
            "Equipo a utilizar": {
                "Equipo a utilizar":"Tina de Acero Inox, Dispersor o Tambo de proceso" 
                },
            "Envase":{
                "Envase_producto":"TAMBOR REUSO AZUL 15-101 P 15-205/CUBETA DE PLASTICO LITOGRAFIADA" ,
                "Envase_control":"Porron de plastico"
                }
            },
        "HUV-128-A":{# Clave
            "Banner 1":"b2",
            "Banner 2":"b8",
            "15-101":{
                "descripcion": "Liquido incoloro; viscosidad Media; libre de particulas",
                "peligro":"m",
                "indicacion":"Pesar y adicionar 15-101 a la tina con una agitación de 550 a 600 r.p.m.",
                "revision":""
                },
            "15-404":{
                "descripcion": "Liquido incoloro; viscosidad Media; libre de particulas",
                "peligro":"tc",
                "indicacion":"Pesar y adicionar 15-404 a la tina",
                "revision":""
                },
            "15-667":{
                "descripcion": "Liquido incoloro; viscosidad Media; libre de particulas",
                "peligro":"a",
                "indicacion":"Pesar y adicionar lentamente 15-667 a la tina",
                "revision":""
                },
            "15-652":{
                "descripcion": "Liquido incoloro; viscosidad Media; libre de particulas",
                "peligro":"cai",
                "indicacion":"Pesar y adicionar lentamente 15-652 a la tina",
                "revision":""
                },
            
            "Equipo a utilizar": {
                "Equipo a utilizar":"Tina de Acero Inox, Dispersor o Tambo de proceso" 
                },
            "Envase":{
                "Envase_producto":"SEGUN ORDEN DE PRODUCCION" ,
                "Envase_control":"Porron de plastico"
                }
            },
        "FAST CURE NATURAL":{ # 21-400
            "Paso 1":{
                "texto":"Antes de iniciar verifique que su equipo se encuentra limpio (tambo ó tina, dispersor, flecha, valvula) VERIFIQUE QUE NO TENGA BASURA NI PARTICULAS"
                },
            "Paso 2":{
                "texto":"Los tambos que utilice deben estar perfectamente escurridos, para lo que si ya los tiene coloquelos boca abajo para que escurran, mientras usted fabrica"
                },
            "Paso 3":{
                "texto":"Si son de 15-101 los puede meter al horno boca abajo para escurrirlos"
                },
            "15-101":{
                "descripcion": "Liquido incoloro; viscosidad Media; libre de particulas",
                "peligro":"m",
                "indicacion":"Pesar el 15-101 y adicionar a la tina o tambo. Inicie agitación colocando su dispersor de 500-550 r.p.m. y verifique que la temperatura suba a 25°C como minimo",
                "revision":""
                },
            "Paso 4":{
                "texto":"MANTENGA DE 500-55O R.P.M. DURANTE TODO SU PROCESO"
                },
            "15-205":{
                "descripcion": "Liquido incoloro; viscosidad Media; libre de particulas",
                "peligro":"tc",
                "indicacion":"Pese el 15-205 y adicione bajo agitación",
                "revision":""
                },
            "15-201":{
                "descripcion": "Liquido incoloro; viscosidad baja con olor frutal; libre de particulas",
                "peligro":"a",
                "indicacion":"Pese el 15-201 y adicionar bajo agitación",
                "revision":""
                },
            "15-200":{
                "descripcion": "Liquido incoloro; viscosidad baja; libre de particulas",
                "peligro":"cai",
                "indicacion":"Pese el 15-200 y adicionar bajo agitación",
                "revision":""
                },
            "25-501":{
                "descripcion": "Dispersión blanca; viscosidad Media; libre de particulas",
                "peligro":"cai",
                "indicacion":"Pese 25-501 y adicione bajo agitacion, deje mezcar durante 5-7 minutos",
                "revision":""
                },
            "Equipo a utilizar": {
                "Equipo a utilizar":"DISPERSOR, TINA Ó TAMBO DE PROCESO, DISCO, VALVULA DE METAL" 
                },
            "Envase":{
                "Envase_producto":"TAMBOR REUSO AZUL 15-101 P 15-205/CUBETA DE PLASTICO LITOGRAFIADA" ,
                "Envase_control":"Porron de plastico"
                }
            },
        "FAST CURE AZUL SUAVE":{ # Clave 21-400-3
            "Paso 1":{
                "texto":"la orden debe estar ajustada quitando el 2.5% del 21-400, en caso de no estar ajustado notifique al jefe de produccion para que le de las indicaciones correspondientes"
                },
            "21-400":{
                "descripcion": "Liquido incoloro; viscosidad Media; libre de particulas",
                "peligro":"m",
                "indicacion":"Pese el 21-400 y adicione a su tina o tambor, Coloque sus r.p.m. de 500-550 y mantenga asi durante todo su proceso",
                "revision":"Antes de  usar el 21-400, ruede el tambor de manera que se homogeinice y pueda proceder a vaciar su tambo o tina de proceso"
                },
            "Paso 2":{
                "texto":"El 21-400 debe estar homogeneo"
                },
            "15-683":{
                "descripcion": "Liquido incoloro; viscosidad Media; libre de particulas",
                "peligro":"tc",
                "indicacion":"Pese el 15-683 y adicione bajo agitación y deje agitar durante 3-5 minutos",
                "revision":""
                },
            "Paso 3":{
                "texto":"Antes de usar el 15-699-A mueva el porron o envase para mezclarlo, este producto debe ser claro y ligeramente amarilo, en caso de estar turbio calientelo a 40°C hasta que se torne claro"
                },
            "15-699-A":{
                "descripcion": "Liquido incoloro; viscosidad Media; libre de particulas",
                "peligro":"a",
                "indicacion":"Si el 15-699-A esta claro pese y y adicionelo bajo agitación",
                "revision":"IMPORTANTE: el 15-699-A debe estar totalmente claro y transparente"
                },
            "29-501":{
                "descripcion": "Liquido incoloro; viscosidad Media; libre de particulas",
                "peligro":"cai",
                "indicacion":"Pese la cantidad indicada en la Ord. Prod. y adicione bajo agitacion, recircule de su tina una cubeta verificando que la s¿mezcla salga homogenea",
                "revision":"Si no sale homogenea, recircule mas material hasta que quede igual"
                },
            "Paso 4":{
                "texto":"NOTIFIQUE AL PLA QUE HARÁ LA IGUALACIÓN PARA QUE LE INDIQUE LA CANTIDAD DE DISPERSIÓN NEGRA QUE LE VAYA A ADICIONAR"
                },
            "29-505":{
                "descripcion": "Dispersion azul; viscosidad Media; libre de particulas",
                "peligro":"cai",
                "indicacion":"Pese la cantidad y adicione de acuerdo a las indicaciones de la persona que iguale",
                "revision":"La dispersion azul debe disolverse por separado, sacando material de su tina y haciendo un masterbatch con su espatula"
                },
            "Paso 5":{
                "texto":"UNA VEZ FINALIZADO LA IGUALACIÓN, PROCEDA A RECIRCULAR NUEVAMENTE MATERIAL PARA VERIFICAR QUE EL COLOR ESTE PAREJO"
                },
            "Paso 6":{
                "texto":"EN CASO DE PRESENTAR DIFERENTES TONOS, RECIRCULE HASTA QUE SALGA DE  UN SOLO TONO Y NOTIFIQUE A LA PERSONA QUE IGUALO PARA QUE VERIFIQUE Y AJUSTE DE SER NECESARIO"
                },
            "15-650":{
                "descripcion": "Liquido incoloro; viscosidad Media; libre de particulas",
                "peligro":"cai",
                "indicacion":"Pese el 15-651-1 y bajo agitacion, adicione, deje mezclando durante 3-5 minutos",
                "revision":""
                },
            "15-651-1":{
                "descripcion": "Liquido incoloro; viscosidad Media; libre de particulas",
                "peligro":"cai",
                "indicacion":"Pese el 15-651-1 y bajo agitacion, adicione, deje mezclando durante 3-5 minutos",
                "revision":""
                },
            "Equipo a utilizar": {
                "Equipo a utilizar":"Tina de Acero Inox, Dispersor o Tambo de proceso, Disco, Valvula de Metal" 
                },
            "Envase":{
                "Envase_producto":"CUBETA DE PLASTICO AMARILLA SIN LITOGRAFIARDE 20KG" ,
                "Envase_control":"Porron de plastico"
                }
            },
        "EPOXINE 500 PRIMER BLANCO A":{ # No. Inv 22-200
            "15-101":{
                "descripcion": "Liquido incoloro; viscosidad Media; libre de particulas",
                "peligro":"m",
                "indicacion":"",
                "revision":""
                },
            "15-222":{
                "descripcion": "Liquido incoloro; viscosidad Media; libre de particulas",
                "peligro":"am",
                "indicacion":"",
                "revision":""
                },
            "15-699-2":{
                "descripcion": "Liquido incoloro; viscosidad Media; libre de particulas",
                "peligro":"0",
                "indicacion":"",
                "revision":""
                },
            "15-683":{
                "descripcion": "Liquido incoloro; viscosidad Media; libre de particulas",
                "peligro":"0",
                "indicacion":"",
                "revision":""
                },
            "15-692":{
                "descripcion": "Dispersion azul; viscosidad Media; libre de particulas",
                "peligro":"ai",
                "indicacion":"",
                "revision":""
                },
            "15-702":{
                "descripcion": "Liquido incoloro; viscosidad Media; libre de particulas",
                "peligro":"0",
                "indicacion":"",
                "revision":""
                },
            "15-698-1":{
                "descripcion": "Liquido incoloro; viscosidad Media; libre de particulas",
                "peligro":"0",
                "indicacion":"",
                "revision":""
                },
            "Equipo a utilizar": {
                "Equipo a utilizar":"Tina de Acero Inox, Dispersor o Tambo de proceso" 
                },
            "Envase":{
                "Envase_producto":"0" ,
                "Envase_control":"Porron de plastico"
                }
            },
        "EPOXINE 500 BLANCO A":{ # No. Inv 22-200
            "Banner 1":"b1",
            "Banner 2":"b2",
            "15-101":{
                "descripcion": "Liquido incoloro; viscosidad Media; libre de particulas",
                "peligro":"m",
                "indicacion":"",
                "revision":""
                },
            "15-214":{
                "descripcion": "Liquido incoloro; viscosidad Media; libre de particulas",
                "peligro":"am",
                "indicacion":"",
                "revision":""
                },
            "15-569":{
                "descripcion": "Liquido incoloro; viscosidad Media; libre de particulas",
                "peligro":"0",
                "indicacion":"",
                "revision":""
                },
            "15-662-1":{
                "descripcion": "Liquido incoloro; viscosidad Media; libre de particulas",
                "peligro":"0",
                "indicacion":"",
                "revision":""
                },
            "15-769":{
                "descripcion": "Dispersion azul; viscosidad Media; libre de particulas",
                "peligro":"ai",
                "indicacion":"",
                "revision":""
                },
            "15-129":{
                "descripcion": "Liquido incoloro; viscosidad Media; libre de particulas",
                "peligro":"0",
                "indicacion":"",
                "revision":""
                },
            "15-617-1":{
                "descripcion": "Liquido incoloro; viscosidad Media; libre de particulas",
                "peligro":"0",
                "indicacion":"",
                "revision":""
                },
            "15-617-1":{
                "descripcion": "Liquido incoloro; viscosidad Media; libre de particulas",
                "peligro":"0",
                "indicacion":"",
                "revision":""
                },
            "15-515":{
                "descripcion": "Liquido incoloro; viscosidad Media; libre de particulas",
                "peligro":"0",
                "indicacion":"",
                "revision":""
                },
            "15-683":{
                "descripcion": "Liquido incoloro; viscosidad Media; libre de particulas",
                "peligro":"0",
                "indicacion":"",
                "revision":""
                },
            "15-690-1":{
                "descripcion": "Liquido incoloro; viscosidad Media; libre de particulas",
                "peligro":"0",
                "indicacion":"",
                "revision":""
                },
            "Equipo a utilizar": {
                "Equipo a utilizar":"Tina de Acero Inox, Dispersor o Tambo de proceso" 
                },
            "Envase":{
                "Envase_producto":"0" ,
                "Envase_control":"Porron de plastico"
                }
            },
        # CONSIDERACIONES A LA HORA DE RECABAR LA INFORMACION
        # SIEMPRE SEGUIR UNA NUMERACION DE LOS PASOS {'PASO 1','xxx'} Y {'PASO 2','xxx'} Y ASI
        "Klipton S-12":{
            "Banner 1":"b1",
            "Banner 2":"b5",
            "15-607":{
                "descripcion": "Liquido incoloro; viscosidad Baja; libre de particulas",
                "peligro":"iat",
                "indicacion":"Pese el 15-607 y adicione a su tina/tambo",
                "revision":"verifique que el liquido este libre de particulas PELIGRO: NO GOLPEAR/AZOTAR RIESGO DE INCENDIO"
                },
            "15-612":{
                "descripcion": "Liquido incoloro; viscosidad Baja; libre de particulas",
                "peligro":"ita",
                "indicacion":"Pese el 15-612 y con plataforma añadalo a la tina/tambo",
                "revision":"verifique que el liquido este libre de particulas PELIGRO: NO GOLPEAR/AZOTAR RIESGO DE INCENDIO"
                },
            "15-690":{
                "descripcion": "Liquido incoloro; viscosidad Baja; libre de particulas",
                "peligro":"iat",
                "indicacion":"Pese el 15-690 y plataforma añadalo a la tina/tambo",
                "revision":"verifique que el liquido este libre de particulas"
                },
            "15-499-3":{
                "descripcion": "Liquido color marron verdoso, viscosidad baja",
                "peligro":"a",
                "indicacion":"Pese el 15-499-3 y plataforma añadalo a la tina/tambo",
                "revision":"verifique que el liquido este libre de particulas"
                },
            "Paso 1":{
                "texto":"COLOQUE SU TINA EN EL DISPERSOR: CUANDO FABRIQUE EN EL DISPERSOR, COLOQUE LA TIERRA FISICA ANTES DE  INICIAR CUALQUIER AGITACION Ó ADICION"
                },
            "15-494":{
                "descripcion": "Polvo blanco de baja densidad",
                "peligro":"0",
                "indicacion":"Pese el 15-494, con plataforma añadalo a la tina/tambo",
                "revision":"Particulas solidas blancas de bajas densidad"
                },
            "15-495":{
                "descripcion": "Polvo blanco de baja densidad",
                "peligro":"0",
                "indicacion":"Pese el 15-495, con plataforma añadalo a la tina/tambo",
                "revision":"Particulas solidas blancas de bajas densidad"
                },
            "15-452-1":{
                "descripcion": "escamas sólidas amarillas",
                "peligro":"0",
                "indicacion":"Pese el 15-452-1 y agite a 600-750 r.p.m.",
                "revision":"verifique que el polvo tenga un color uniforme, sin grumos ni este apelmazado"
                },
            "Paso 2":{
                "texto":"AGITE POR UN TIEMPO DE 60-80 MINUTOS. AL FINALIZAR VERIFIQUE QUE NO HAYA GRUMOS, PARA LO CUAL DETENGA LA AGITACION Y ESPERE A VER SI NO FLOTAN GRUMOS"
                },
            "15-428-1":{
                "descripcion": "Polvo amarillento de baja densidad",
                "peligro":"ia",
                "indicacion":"Pese el 15-428-1 y adicione bajo agitacion",
                "revision":"verifique que el polvo tenga un color uniforme, sin grumos ni este apelmazado"
                },
            "Paso 3":{
                "texto":"AGITE DE 25-30 MINUTOS"
                },
            "15-537":{
                "descripcion": "Polvo blanco ligero",
                "peligro":"t",
                "indicacion":"Pese el 15-537 y adicione bajo agitacion",
                "revision":"verifique que el polvo tenga un color uniforme, sin grumos ni este apelmazado"
                },
            "Paso 4":{
                "texto":"AGITE DE 10-12 MINUTOS"
                },
            "15-562":{
                "descripcion": "Polvo Blanco",
                "peligro":"t",
                "indicacion":"Pese el 15-562 y adicione bajo agitacion, agitar durante 10-12 minutos y recircule 2 cubetas",
                "revision":"verifique que el polvo tenga un color uniforme, sin grumos ni este apelmazado"
                },
            "Paso 5":{
                "texto":"AGITE DE 10-12 MINUTOS; EN LA MAQUINA CHINA DETENGA LA AGITACION Y RECIRCULE 2 CUBETAS"
                },
            "15-515":{
                "descripcion": "Polvo blanco muy ligero",
                "peligro":"t",
                "indicacion":"Pese el 15-515, divida en 4 partes",
                "revision":"verifique que el saco no este roto"
                },
            "Paso 6":{
                "texto":"SIN AGITACION ADICIONE EN 4 PARTES EL 15-515 Y TAPE LA TINA. POSTERIORMENTE INICIE AGITACION."
                },
            "Banner 3":"b4",
            "15-610":{
                "descripcion": "Liquido incoloro; viscosidad Baja; libre de particulas",
                "peligro":"ia",
                "indicacion":"Pese el 15-610, agite de 5 a 7 minutos y recircule",
                "revision":"verifique que el liquido este libre de particulas"
                },
            "Paso 7":{
                "texto":"EN LA MAQUINA CHINA RECIRCULE 2 CUBETAS"
                },
            "Equipo a utilizar": {
                "Equipo a utilizar":"Dispersor (J o I) y Tina/ Maquina China" 
                },
            "Envase":{
                "Envase_producto":"CUBETA DE PLASTICO LITOGRAFIADA 20 KG" ,
                "Envase_control":"Porron de plastico"
                }
            },
        "Klipton S-11":{
            "Banner 1":"b1",
            "Banner 2":"b2",
            "Paso 1":{
                "texto":"ADVERTENCIA: el 15-607 Y EL 15-612 SON PRODUCTOS ALTAMENTE INFLAMABLES, POR LO QUE EVITE AZOTAR LOS TAMBOS Ó GOLPEAR LAS ENTRADAS, YA QUE ESTO PUEDE PROVOCAR CHISPAS Y GENERAR UNA EXPLOSION"
                },
            "15-607":{
                "descripcion": "Liquido incoloro; viscosidad Baja; libre de particulas",
                "peligro":"iat",
                "indicacion":"Pese el 15-607 y y adicione a su tina/tambo",
                "revision":"verifique que el liquido este libre de particulas PELIGRO: NO GOLPEAR/AZOTAR RIESGO DE INCENDIO"
                },
            "15-612":{
                "descripcion": "Liquido incoloro; viscosidad Baja; libre de particulas",
                "peligro":"it",
                "indicacion":"Pese el 15-612 y con polipasto/plataforma añadalo a la tina/tambo",
                "revision":"verifique que el liquido este libre de particulas PELIGRO: NO GOLPEAR/AZOTAR RIESGO DE INCENDIO"
                },
            "15-690":{
                "descripcion": "Liquido incoloro; viscosidad Baja; libre de particulas",
                "peligro":"iat",
                "indicacion":"Pese el 15-690 y con polipasto/plataforma añadalo a la tina/tambo",
                "revision":"verifique que el liquido este libre de particulas"
                },
            "Paso 2":{
                "texto":"COLOQUE SU TINA EN EL DISPERSOR: CUANDO FABRIQUE EN EL DISPERSOR, COLOQUE LA TIERRA FISICA ANTES DE  INICIAR CUALQUIER AGITACION Ó ADICION; TAPE PERFECTAMENTE SU TINA PARA EVITAR EVAPORACION"
                },
            "15-494":{
                "descripcion": "Polvo blanco de baja densidad",
                "peligro":"0",
                "indicacion":"Pese el 15-494, con plataforma añadalo a la tina/tambo",
                "revision":"Particulas solidas blancas de bajas densidad"
                },
            "15-495":{
                "descripcion": "Polvo blanco de baja densidad",
                "peligro":"0",
                "indicacion":"Pese el 15-495, con plataforma añadalo a la tina/tambo",
                "revision":"Particulas solidas blancas de bajas densidad"
                },
            "15-499-3":{
                "descripcion": "Liquido marron verdoso poco viscoso",
                "peligro":"a",
                "indicacion":"Pese el 15-499-3 y con polipasto/plataforma añadalo a la tina/tambo",
                "revision":"verifique que el liquido este libre de particulas"
                },
            "Paso 3":{
                "texto":"AGITE POR UN TIEMPO DE 60-80 MINUTOS. AL FINALIZAR VERIFIQUE QUE NO HAYA GRUMOS, PARA LO CUAL DETENGA LA AGITACION Y ESPERE A VER SI NO FLOTAN GRUMOS"
                },
            "15-452-1":{
                "descripcion": "Polvo amarillento",
                "peligro":"0",
                "indicacion":"Pese el 15-452-1 y agite a 600-750 r.p.m.",
                "revision":"verifique que el polvo tenga un color uniforme, sin grumos ni este apelmazado"
                },
            "15-428-1":{
                "descripcion": "Polvo amarillento",
                "peligro":"ia",
                "indicacion":"Pese el 15-428-1 y adicione bajo agitacion",
                "revision":"verifique que el polvo venga en pequeños granulos, caso contrario, triture antes de adicionar"
                },
            "Paso 4":{
                "texto":"DEJE AGITAR DE 25-30 MINUTOS"
                },
            "15-537":{
                "descripcion": "Carga color blanco",
                "peligro":"t",
                "indicacion":"Pese el 15-537 y adicione bajo agitacion",
                "revision":"verifique que el polvo tenga un color uniforme, sin grumos ni este apelmazado"
                },
            "Paso 5":{
                "texto":"AGITE DE 10-12 MINUTOS"
                },
            "15-562":{
                "descripcion": "Polvo Blanco",
                "peligro":"t",
                "indicacion":"Pese el 15-562 y adicione bajo agitacion, agitar durante 10-12 minutos y recircule 2 cubetas",
                "revision":"verifique que el polvo tenga un color uniforme, sin grumos ni este apelmazado"
                },
            "Paso 6":{
                "texto":"AGITE DE 10-12 MINUTOS; EN LA MAQUINA CHINA DETENGA LA AGITACION Y RECIRCULE 2 CUBETAS"
                },
            "15-515":{
                "descripcion": "Polvo",
                "peligro":"0",
                "indicacion":"Pese el 15-515, divida en 4 partes",
                "revision":"verifique que el polvo tenga un color uniforme, sin grumos ni este apelmazado"
                },
            "Paso 7":{
                "texto":">1ra Parte: SIN AGITACION ADICIONE LA 1RA 1/4 PARTE. INICIE AGITACION UNA VEZ INCORPORADO DETENGA LA AGITACION"
                },
            "15-610":{
                "descripcion": "Liquido incoloro; viscosidad Baja; libre de particulas",
                "peligro":"ia",
                "indicacion":"Pese el 15-610, agite de 5 a 7 minutos y recircule",
                "revision":"verifique que el liquido este libre de particulas"
                },
            "Paso 8":{
                "texto":"EN LA MAQUINA CHINA RECIRCULE 2 CUBETAS"
                },
            "Equipo a utilizar": {
                "Equipo a utilizar":"Dispersor (J o I) y Tina/ Maquina China" 
                },
            "Envase":{
                "Envase_producto":"CUBETA DE PLASTICO LITOGRAFIADA 20 KG" ,
                "Envase_control":"Porron de plastico"
                }
            },
        "Klipton S-15":{
            "Banner 1":"b8",
            "Banner 2":"b2",
            "Banner 3":"b5",
            "15-699-E":{
                "descripcion": "Liquido Amarillo verdoso; viscosidad Baja; libre de particulas",
                "peligro":"it",
                "indicacion":"Pese el 15-699-E y con plataforma añadalo a la tina/tambo, sin agitar",
                "revision":"verifique que el liquido este libre de particulas PELIGRO: NO GOLPEAR/AZOTAR RIESGO DE INCENDIO"
                },
            "15-607":{
                "descripcion": "Liquido incoloro; viscosidad Baja; libre de particulas",
                "peligro":"iat",
                "indicacion":"Pese el 15-607 y y adicione a su tina/tambo, sin agitar",
                "revision":"verifique que el liquido este libre de particulas PELIGRO: NO GOLPEAR/AZOTAR RIESGO DE INCENDIO"
                },
            "15-690":{
                "descripcion": "Liquido incoloro; viscosidad Baja; libre de particulas",
                "peligro":"iat",
                "indicacion":"Pese el 15-690 y inicie la Agitación a 500-550 R.P.M.",
                "revision":"verifique que el liquido este libre de particulas"
                },
            "15-499-3":{
                "descripcion": "Liquido marron verdoso poco viscoso",
                "peligro":"a",
                "indicacion":"Pese el 15-499-3 y agregue bajo agitacion de 650-700 R.P.M.",
                "revision":"verifique que el liquido este libre de particulas"
                },
            "Paso 1":{
                "texto":"COLOQUE SU TINA EN EL DISPERSOR: CUANDO FABRIQUE EN EL DISPERSOR, COLOQUE LA TIERRA FISICA ANTES DE  INICIAR CUALQUIER AGITACION Ó ADICION; TAPE PERFECTAMENTE SU TINA PARA EVITAR EVAPORACION"
                },
            "15-494":{
                "descripcion": "Polvo blanco de baja densidad",
                "peligro":"0",
                "indicacion":"Pese el 15-494, y agregue a la mezcla bajo agitación de 500-550 R.P.M.",
                "revision":"Particulas solidas blancas de bajas densidad"
                },
            "15-495":{
                "descripcion": "Polvo blanco de baja densidad",
                "peligro":"0",
                "indicacion":"Pese el 15-495, y agregue a la mezcla bajo agitación de 500-550 R.P.M.",
                "revision":"Particulas solidas blancas de bajas densidad"
                },
            "Paso 2":{
                "texto":"AGITE POR UN TIEMPO DE 60-80 MINUTOS. AL FINALIZAR VERIFIQUE QUE NO HAYA GRUMOS, PARA LO CUAL DETENGA LA AGITACION Y CON UN ESPATULA CHEQUE SI NO HAY GRUMOS"
                },
            "15-452-1":{
                "descripcion": "Polvo amarillento",
                "peligro":"0",
                "indicacion":"Pese el 15-452-1 y agreguelo manteniendo la agitacion (650-700 R.P.M.)",
                "revision":"verifique que el polvo tenga un color uniforme, sin grumos ni este apelmazado"
                },
            "15-428-1":{
                "descripcion": "Polvo amarillento",
                "peligro":"ia",
                "indicacion":"Pese el 15-428-1 y adicione bajo agitacion",
                "revision":"verifique que el polvo venga en pequeños granulos, caso contrario, triture antes de adicionar"
                },
            "Paso 3":{
                "texto":"DEJE AGITAR DE 25-30 MINUTOS"
                },
            "15-537":{
                "descripcion": "Carga color blanco",
                "peligro":"t",
                "indicacion":"Pese el 15-537 y agreguelo a la tina. Aumente las R.P.M. a 750 800 R.P.M",
                "revision":"verifique que el polvo tenga un color uniforme, sin grumos ni este apelmazado"
                },
            "Paso 4":{
                "texto":"AGITE DE 10-12 MINUTOS"
                },
            "15-562":{
                "descripcion": "Polvo Blanco",
                "peligro":"t",
                "indicacion":"Pese el 15-562 y adicione bajo agitacion, agitar durante 10-12 minutos y recircule 2 cubetas",
                "revision":"verifique que el polvo tenga un color uniforme, sin grumos ni este apelmazado"
                },
            "Paso 5":{
                "texto":"AGITE DE 10-12 MINUTOS; EN EL DISPERSOR DETENGA LA AGITACION"
                },
            "15-515":{
                "descripcion": "Polvo blanco fino ligero",
                "peligro":"0",
                "indicacion":"Pese el 15-515, divida en 4 partes",
                "revision":"verifique que el polvo tenga un color uniforme, sin grumos ni este apelmazado"
                },
            "Banner 4":"b4",
            "Equipo a utilizar": {
                "Equipo a utilizar":"Dispersor (J o I) y Tina" 
                },
            "Envase":{
                "Envase_producto":"CUBETA DE PLASTICO LITOGRAFIADA 20 KG" ,
                "Envase_control":"Lata de 1/2 Kilo"
                }
            },
        "Klipton 612":{ # 23-205
            "Banner 1":"b2",
            "15-332":{
                "descripcion": "Liquido incoloro o amarillo; viscosidad Media; libre de particulas",
                "peligro":"a",
                "indicacion":"Pese el 15-332 y y adicione al reactor",
                "revision":"verifique que el liquido este libre de particulas y sea claro"
                },
            "Paso 1":{
                "texto":"Tip: Cargue en el reactor, una cantidad superior a la necesaria de 15-332, y por la parte inferior, extraiga el excedente"
                },
            "Paso 2":{
                "texto":"Una vez cargado la cantidad necesaria de 15-332 en el reactor, conecte el Nitrogeno al reactor por la escotilla"
                },
            "15-335":{
                "descripcion": "Liquido incoloro; viscosidad Media; libre de particulas",
                "peligro":"tic",
                "indicacion":"Pese el 15-335 y con polipasto/plataforma añadalo al reactor",
                "revision":"verifique que el liquido este libre de particulas "
                },
            "Paso 3":{
                "texto":"Una vez cargado la cantidad necesaria de 15-335, manteniendo el nitrogeno, deje agitar entre 5 a 15 minutos, debe controlar que la temperatura no suba arriba de 45°C"
                },
            "Paso 3":{
                "texto":"La mezcla al inicio se vera liquida, mucho oleaje, a medida que avanza la agitacion la mezcla toma consistencia de resina, una vez que no haya oleaje manteniendo la agitacion, es momento de añadir el siguiente reactivo"
                },
            "15-310":{
                "descripcion": "Liquido incoloro; viscosidad Baja; libre de particulas",
                "peligro":"c",
                "indicacion":"Pese el 15-310 y añadalo a reactor por la escotilla del mismo, manteniedo el nitrogeno y la temperatura",
                "revision":"verifique que el liquido este libre de particulas"
                },
            "Equipo a utilizar": {
                "Equipo a utilizar":"REACTOR, NITROGENO, POLIPASTO, PLATAFORMA" 
                },
            "Envase":{
                "Envase_producto":"TAMBOR" ,
                "Envase_control":"Porron de plastico"
                }
            },
        "Bullton 5085":{ # 23-110 o 25-110
            "Banner 1":"b7",
            "Banner 2":"b6",
            "Banner 3":"b2",
            "Paso 1":{
                "texto":"Verifica) Que los tambores del 15-443, 15-470 y 15-416 se encuentren cerrados de fabricacion, en caso de que el tambo se haya usado, consultar al SPR si se va a deshidratar "
                },
            "Paso 2":{
                "texto":"Paso 1) Cargue del 15-443, 15-470 y 15-416 al reactor. Para verificar el color claro de sus materias primas, utilice cubetas blancas, las cuales deben estar perfectamente limpias y secas"
                },
            "15-443":{
                "descripcion": "Liquido incoloro o amarillo; viscosidad Media; libre de particulas",
                "peligro":"a",
                "indicacion":"Pese el 15-443 y adicione al reactor",
                "revision":"verifique que el liquido este libre de particulas y sea totalmente claro"
                },
            "15-470":{
                "descripcion": "Liquido incoloro; viscosidad Media; libre de particulas",
                "peligro":"",
                "indicacion":"Pese el 15-470 y con polipasto-plataforma añadalo al reactor",
                "revision":"verifique que el liquido este libre de particulas y sea totalmente claro"
                },
            "15-416":{
                "descripcion": "Liquido incoloro; viscosidad Baja; libre de particulas",
                "peligro":"",
                "indicacion":"Pese el 15-416 y añadalo al reactor, Cierre el reactor y suba la temperatura hasta 76°C±2°C",
                "revision":""
                },
            "Paso 3":{
                "texto":"Paso 2) Deshidratación: En el  reactor se agitan y se lleva a cabo la deshidratación a 115°C±5°C y 50 cm Hg de vacío durante 30 minutos y hasta 1 hora en clima humedo/lluvioso"
                },
            "Paso 4":{
                "texto":"Paso 3) Una vez concluida la deshidratación bajar la temperatura hasta 76±2°C antes de hacer las adiciones del 15-351."
                },
            "Paso 5":{
                "texto":"Paso 4) Si despues del proceso de fabricación, sobra material en los tambores, ponga una capa de nitrogeno para evitar la entrada de humedad."
                },
            "15-351":{
                "descripcion": "Liquido ligeramente amarillo; viscosidad Baja; libre de particulas",
                "peligro":"ar",
                "indicacion":"De la cantidad de este reactivo, dividalo en diez partes, y agregue cada decima parte en intervalos de 5 minutos usando una cubeta limpia y seca",
                "revision":"verifique que el liquido este libre de particulas, y transparente."
                },
            "Paso 6":{
                "texto":"Si el 15-351 está solidificado o cristalizado, no lo use. Cuidar que durante la adición de temperatura sea menor a 84°C±2°C, durante la adicion del 15-351 ponga una capa de nitrogeno para evitar la entrada de humedad"
                },
            "Paso 7":{
                "texto":"Una vez terminada la adición, espere 15 minutos a que estabilice la temperatura de la mezcla a 98±2°C bajo agitación constante. Es importante que mantenga la temperatura a 98±2°C al realizar el cocinado para obtener un %NCO deseado"
                },
            "25-501":{
                "descripcion": "Liquido incoloro; viscosidad Baja; libre de particulas",
                "peligro":"ai",
                "indicacion":"Quince minutos antes de terminar el cocinado, pese el 25-501, y adicione al reactor por vacio a 50cm Hg",
                "revision":"verifique que el liquido este libre de particulas"
                },
            "Paso 8":{
                "texto":"Al tomar una muestra, primero saque aprox 10 Kg de producto en una cubeta limpia y seca, posteriormente tome su muestra en una lata limpia y seca"
                },
            "Paso 9":{
                "texto":"El Bullton 5085 es un material muy sensible a la humedad, por lo que se recomienda al momento de envasar, poner una capa de nitrogeno sobre la superficie del prepolimero y cerrar perfectamente el envase."
                },
            "Equipo a utilizar": {
                "Equipo a utilizar":"REACTOR, POLIPASTO" 
                },
            "Envase":{
                "Envase_producto":"TAMBOR NUEVO o EXCLUSIVAMENTE de 15-351, no usar ningun otro, solo los 2 anteriores" ,
                "Envase_control":"Porron de plastico"
                }
            },
        "Bullton 2013":{ # 25-105
            "Banner 1":"b2", # BANNER DE  EQUIPO DE PROTECCION
            "Banner 2":"b8", # MANTENER TU REACTOR LIMPIO
            "Banner 3":"b7", # BANNER DE CONSIDERACIONES ANTES DE USAR E PROCESO
            "Paso 1":{
                "texto":"Paso 1) Fundición del 15-415: Se deja por 12 horas el tambo del 15-415 a una temperatura de 105±5°C, se saca del horno, se abre uno de los tapones para liberar presión"
                },
            "Paso 2":{
                "texto":"Paso 2) Adicion del 15-415 al reactor: se pesa y se carga al reactor por medio de uso del vacio la cantidad de 15-415 requerida segun la Orden de Producción"
                },
            "15-415":{
                "descripcion": "Sólido/Liquido Fundido incoloro o amarillo; viscosidad Alta-Media; libre de particulas",
                "peligro":"a",
                "indicacion":"Cuidado!: Tambor Caliente, Pese el tambor lleno y succione el peso necesario para la fabricacion del producto",
                "revision":"verifique que el liquido este fundido y tenga un color claro. Notifique al SPR si no es asi"
                },
            "Paso 3":{
                "texto":"Paso 3) Deshidratacion del 15-415: Una vez el 15-415 de forma liquida en el reactor, se cierra, se calienta a 110-115°C y con agitación, se aplica un vacio de 30-60 cmHg durante 30 minutos"
                },
            "Paso 4":{
                "texto":"Paso 4) Nitrogeno en los tambores: Mientras se lleva a cabo la deshidratación, coloque una capa de Nitrogeno en/los tambores de 15-415, para su posterior uso"
                },
            "Paso 5":{
                "texto":"Consejo) Para verificar el color claro de sus materias primas, utilice cubetas blancas, las cuales deben estar perfectamente limpias y secas. Si usas envases extras, siempre deben estar perfectamente limpios y secos"
                },
            "Paso 6":{
                "texto":"Paso 5) Enfriamiento del 15-415: Una vez terminada la deshidratación, se enfria el reactor con agua hasta llegar a una temperatura de 48±2°C"
                },
            "Paso 7":{
                "texto":"Paso 6) Adiciones del 15-351: Una vez alcanzada la temperatura estable, pese la cantidad del 15-351, y haga adiciones cada 5 minutos en un tiempo total de entre 45 y 50 minutos. Cuide en este punto de que la temperatura marcada no exceda los 55°C. Recomendacion: a los 53°C agregue agua de enfriamiento"
                },
            "15-351":{
                "descripcion": "Liquido ligeramente amarillo; viscosidad Baja",
                "peligro":"ar",
                "indicacion":"Haga adiciones de 15-351 y que no exceda la temperatura de 53±2°C",
                "revision":"verifique que el liquido este libre de particulas, y transparente."
                },
            "Paso 8":{
                "texto":"Paso 7) Cocinado: Una vez terminada la adición, espere 15 minutos a que estabilice, ahora la temperatura de la mezcla debe ser 83±2°C bajo agitación constante. Es importante que mantenga la temperatura a 83(+/-)2°C al realizar el cocinado para obtener un %NCO deseado: Duración del cocinado: 2 horas"
                },
            "Paso 9":{
                "texto":"Paso 8) Adicion del 25-501: 15 minutos antes de terminar las 2 horas del cocinado, agregue lo indicado de 25-501 con vacío de 50 cmHG segun la Orden de Producción "
                },
            "25-501":{
                "descripcion": "Liquido incoloro; viscosidad Baja; libre de particulas",
                "peligro":"ai",
                "indicacion":"Quince minutos antes de terminar el cocinado, pese el 25-501 y adicione al reactor",
                "revision":""
                },
            "Paso 10":{
                "texto":"El Bullton 2013 es un material muy sensible a la humedad, por lo que se recomienda al momento de envasar, poner una capa de nitrogeno sobre la superficie del prepolimero y cerrar perfectamente el envase."
                },
            "Paso 11":{
                "texto":"La temperatura de envasado es de 60±2°C"
                },
            "Equipo a utilizar": {
                "Equipo a utilizar":"REACTOR, POLIPASTO" 
                },
            "Envase":{
                "Envase_producto":"TAMBOR NUEVO o EXCLUSIVAMENTE de 15-315, no usar ningun otro, solo los 2 anteriores" ,
                "Envase_control":"Lata de 1/2"
                # PARA ESTE PRODUCTO UNICAMENTE
                # El bullton 5085 es un material muy sensible a la humedad por lo que se recomienda al momento de envasar, poner una capa de nitrogeno sobre la superficie del prepolimero y cerrar perfectamente el envase. La temperatura de envasado es de 60°C±2°C
                }
             },
        "Klipton C-35":{ # 23-116
            "Banner 1":"b2", # BANNER DE  EQUIPO DE PROTECCION
            "Banner 2":"b7", # BANNER DE CONSIDERACIONES ANTES DE USAR E PROCESO
            "Banner 3":'b5', # BANNER DE advert. de explosion
            "Banner 4":'b1', # BANNER DE PREPARACION
            "Paso 1":{
                "texto":"Cualquier desviación de cantidades de MP que no correspondan a lo indicado en la orden de producción debe ser reportado al Jefe de Producción, el cual indicará el seguimiento para el producto"
                },
            "Paso 2":{
                "texto":"No adicionar ningun producto ajeno a los indicados en la orden de producción"
                },
            "15-607":{
##                "nombre_quimico": "xxxxx",
                "descripcion": "Liquido incoloro; viscosidad Baja; libre de particulas",
                "peligro":"iat",
                "indicacion":"Pese el 15-607 y adicione a su tina/tambo",
                "revision":"verifique que el liquido este libre de particulas PELIGRO: NO GOLPEAR/AZOTAR RIESGO DE INCENDIO"
##                "viscosidad":"xxxcp"
                },
            "15-612":{
##                "nombre_quimico": "xxxxx",        
                "descripcion": "Liquido incoloro; viscosidad Baja; libre de particulas",
                "peligro":"ita",
                "indicacion":"Pese el 15-612 y con plataforma añadalo a la tina/tambo",
                "revision":"verifique que el liquido este libre de particulas PELIGRO: NO GOLPEAR/AZOTAR RIESGO DE INCENDIO"
##                "viscosidad":"xxxcp"
                },
            "Paso 3":{
                "texto":"Coloque su tina en el dispersor. Cuando fabrique, coloque la tierra fisica antes de cualquier agitación ó adición"
                },
            "15-494":{
##                "nombre_quimico": "xxxxx",
                "descripcion": "Polvo blanco de baja densidad",
                "peligro":"0",
                "indicacion":"Pese el 15-494 y adicione bajo agitación",
                "revision":"Coloque su dispersor a 650-700 r.p.m.",
##                "viscosidad":""
                },
            "15-495":{
##                "nombre_quimico": "xxxxx",
                "descripcion": "Polvo blanco de baja densidad",
                "peligro":"0",
                "indicacion":"Pese y adicione el 15-495 bajo agitación.",
                "revision":"Agite por durante 15-20 minutos",                
##                "viscosidad":""
                },
            "15-428-1":{
##                "nombre_quimico": "xxxxx",
                "descripcion": "Polvo amarillento",
                "peligro":"ia",
                "indicacion":"Pese el 15-428-1 y adicione bajo agitacion",
                "revision":"verifique que el polvo venga en pequeños granulos, caso contrario, triture antes de adicionar",
##                "viscosidad":""
                },
            "Paso 4":{
                "texto":"Agite por un tiempo de 25-30 minutos. Al finalizar, verifique que no haya grumos, para lo cual detenga la agitación y espere a ver si no flotan los grumos"
                },
            "15-452-1":{
##                "nombre_quimico": "xxxxx",
                "descripcion": "Polvo amarillento",
                "peligro":"0",
                "indicacion":"Pese el 15-452-1 y agite a 600-750 r.p.m.",
                "revision":"verifique que el polvo tenga un color uniforme, sin grumos ni este apelmazado",
##                "viscosidad":""
                },
            "Paso 4":{
                "texto":"Agite por un tiempo de 25-30 minutos."
                },            
            "15-515":{
##                "nombre_quimico": "xxxxx",
                "descripcion": "Polvo blanco muy ligero",
                "peligro":"t",
                "indicacion":"Pese el 15-515, divida en 4 partes",
                "revision":"",
##                "viscosidad":""
                },
            "Banner 5":'b4', # BANNER DE "MEZCLAR EN 4 PARTES" 
            "15-610":{
##                "nombre_quimico": "xxxxx",
                "descripcion": "Liquido incoloro; viscosidad Baja; libre de particulas",
                "peligro":"ia",
                "indicacion":"Pese el 15-610, agite de 5 a 7 minutos y recircule",
                "revision":"verifique que el liquido este libre de particulas",
##                "viscosidad":"xxx cp"
                },
            "Paso 5":{
                "texto":"Si recircula, utilice una cubeta limpia y seca."
                },     
            "Equipo a utilizar": {
                "Equipo a utilizar":"DISPERSOR (J o I), TINA Y MAQUINA CHINA" 
                },
            "Envase":{
                # SE FILTRA CON MALLA? R: si, malla de 150 micras para tina. En reactor, malla de punta
                "Envase_producto":"CUBETA AMARILLA LITOGRAFIADA" ,
                "Envase_control":"Lata de 1/2 L"
                }
            },
        "AC-4368":{ # 20-100
            "Banner 1":"b2",
            "15-332":{
                "descripcion": "Liquido incoloro o amarillo; viscosidad Media; libre de particulas",
                "peligro":"a",
                "indicacion":"Pese el 15-332 y y adicione al reactor",
                "revision":"verifique que el liquido este libre de particulas y sea claro"
                },
            "Paso 1":{
                "texto":"Tip: Cargue en el reactor, una cantidad superior a la necesaria de 15-332, y por la parte inferior, extraiga el excedente"
                },
            "Paso 2":{
                "texto":"Una vez cargado la cantidad necesaria de 15-332 en el reactor, conecte el Nitrogeno al reactor por la escotilla"
                },
            "15-335":{
                "descripcion": "Liquido incoloro; viscosidad Media; libre de particulas",
                "peligro":"tic",
                "indicacion":"Pese el 15-335 y con polipasto/plataforma añadalo al reactor",
                "revision":"verifique que el liquido este libre de particulas "
                },
            "Paso 3":{
                "texto":"Una vez cargado la cantidad necesaria de 15-335, manteniendo el nitrogeno, deje agitar entre 5 a 15 minutos, debe controlar que la temperatura no suba arriba de 45°C"
                },
            "Paso 3":{
                "texto":"La mezcla al inicio se vera liquida, mucho oleaje, a medida que avanza la agitacion la mezcla toma consistencia de resina, una vez que no haya oleaje manteniendo la agitacion, es momento de añadir el siguiente reactivo"
                },
            "15-310":{
                "descripcion": "Liquido incoloro; viscosidad Baja; libre de particulas",
                "peligro":"c",
                "indicacion":"Pese el 15-310 y añadalo a reactor por la escotilla del mismo, manteniedo el nitrogeno y la temperatura",
                "revision":"verifique que el liquido este libre de particulas"
                },
            "Equipo a utilizar": {
                "Equipo a utilizar":"REACTOR, NITROGENO, POLIPASTO, PLATAFORMA" 
                },
                # SE FILTRA CON MALLA? R:
            "Envase":{
                "Envase_producto":"TAMBOR" ,
                "Envase_control":"Porron de plastico"
                }
            },
        "AC-74":{ # 20-109
            "Banner 1":"b2",
            "15-332":{
                "descripcion": "Liquido incoloro o amarillo; viscosidad Media; libre de particulas",
                "peligro":"a",
                "indicacion":"Pese el 15-332 y y adicione al reactor",
                "revision":"verifique que el liquido este libre de particulas y sea claro"
                },
            "Paso 1":{
                "texto":"Tip: Cargue en el reactor, una cantidad superior a la necesaria de 15-332, y por la parte inferior, extraiga el excedente"
                },
            "Paso 2":{
                "texto":"Una vez cargado la cantidad necesaria de 15-332 en el reactor, conecte el Nitrogeno al reactor por la escotilla"
                },
            "15-335":{
                "descripcion": "Liquido incoloro; viscosidad Media; libre de particulas",
                "peligro":"tic",
                "indicacion":"Pese el 15-335 y con polipasto/plataforma añadalo al reactor",
                "revision":"verifique que el liquido este libre de particulas "
                },
            "Paso 3":{
                "texto":"Una vez cargado la cantidad necesaria de 15-335, manteniendo el nitrogeno, deje agitar entre 5 a 15 minutos, debe controlar que la temperatura no suba arriba de 45°C"
                },
            "Paso 3":{
                "texto":"La mezcla al inicio se vera liquida, mucho oleaje, a medida que avanza la agitacion la mezcla toma consistencia de resina, una vez que no haya oleaje manteniendo la agitacion, es momento de añadir el siguiente reactivo"
                },
            "15-310":{
                "descripcion": "Liquido incoloro; viscosidad Baja; libre de particulas",
                "peligro":"c",
                "indicacion":"Pese el 15-310 y añadalo a reactor por la escotilla del mismo, manteniedo el nitrogeno y la temperatura",
                "revision":"verifique que el liquido este libre de particulas"
                },
            "Equipo a utilizar": {
                "Equipo a utilizar":"REACTOR, NITROGENO, POLIPASTO, PLATAFORMA" 
                },
            "Envase":{
                # SE FILTRA CON MALLA? R:
                "Envase_producto":"TAMBOR" ,
                "Envase_control":"Porron de plastico"
                }
            },
        "Dispersion Negra": { # PRODUCTO 1 # Clave 29-506
            "Banner 1":"b2",# BANNER DE  EQUIPO DE PROTECCION
            "Banner 2":"b12",# MANTENGA LIMPIO SU REACTOR 
            "Paso 1":{
                "texto": "Sacar previamente del horno el 15-774 por 1.5 horas" 
                },
            "15-101":{
                "descripcion": "Liquido incoloro/amarillento; viscosidad Alta; libre de particulas",
                "peligro":"m",
                "indicacion":"Pese el 15-101 y adicione a la tina de dispersión sin agitar",
                "revision":"Revise que la resina sea clara y libre de particulas, evite calentar la resina"
                },
            "15-699-10":{
                "descripcion": "Liquido amarillento; viscosidad Alta; libre de particulas",
                "peligro":"0",
                "indicacion":"Pese el 15-699-10, y añadalo bajo agitación",
                "revision":"verifique que el liquido este libre de particulas"
                },
            "Paso 2":{
                "texto": "Durante 3 minutos, mezcle el 15-101 y 15-699-10" 
                },
            "15-774":{
                "descripcion": "Liquido incoloro; viscosidad Baja; libre de particulas",
                "peligro":"m",
                "indicacion":"Pese el 15-340-1 y añadalo con ayuda de un Cucharon, poco a poco",
                "revision":"IMPORTANTE: EVITAR QUE LA MEZCLA SOBREPASE LOS 50-55°C"
                },
            "Paso 3":{
                "texto":"Solicitar Medición de Fineza a los 10 Minutos despues de la adicion total de 15-772"
                },
            "Banner 3":"b10", # Solicitar Fineza
            "15-633":{
                "descripcion": "Liquido incoloro; viscosidad Baja; libre de particulas",
                "peligro":"0",
                "indicacion":"Pese el 15-633 y agreguelo una vez la mezcla se haya vuelto muy viscosa o pesada",
                "revision":"verifique que el liquido este libre de particulas"
                },
            "Paso 4":{
                "texto":"Dejar agitar durante 5 minutos hasta que la mezcla se vea homogenea"
                },
            "15-417":{
                "descripcion": "Liquido incoloro; viscosidad Baja; libre de particulas",
                "peligro":"t",
                "indicacion":"Pese el 15-417 y agregue bajo agitación. Deje agitar durante 10 minutos",
                "revision":"verifique que el liquido este libre de particulas"
                },
            "Banner 4":"b8",
            "Paso 4":{
                "texto":"Adicionar el 15-417 bajo agitación, agite durante 10 minutos"
                },
            "Equipo a utilizar": {
                "Equipo a utilizar":"Tina de Acero Inox, Dispersor o Tambo de proceso" 
                },
            "Envase":{
                "Envase_producto":"Envase indicado en la Orden de Producción" ,
                "Envase_control":"Porron de plastico"
                }
            },
        "TC9": { # PRODUCTO 1 # Clave 24-803
            "Banner 1":"b2",# BANNER DE  EQUIPO DE PROTECCION
            "Banner 2":"b12",# MANTENGA LIMPIO SU REACTOR 
            "Paso 1":{
                "texto": "Agregue el 15-101 y el 15-172 a su tina o tambor de proceso sin agitación" 
                },
            "15-101":{
                "descripcion": "Liquido incoloro/amarillento; viscosidad Alta; libre de particulas",
                "peligro":"am",
                "indicacion":"",
                "revision":"Revise que la resina sea clara y libre de particulas, evite calentar la resina"
                },
            "15-172":{        
                "descripcion": "Liquido transparente claro, viscosidad media",
                "peligro":"am",
                "indicacion":"",
                "revision":"Revise que la resina sea clara y libre de particulas, evite calentar la resina"
                },
            "29-501":{        
                "descripcion": "Dispersión Blanca",
                "peligro":"am",
                "indicacion":"Pese la cantidad de 29-501",
                "revision":""
                },
            "29-505":{        
                "descripcion": "Dispersión Azul",
                "peligro":"am",
                "indicacion":"Pese la cantidad de 29-505",
                "revision":""
                },
            "Paso 2":{
                "texto": "Una vez pesadas las cantidades, inicie agitación y bajo agitacion agregue las dispersiónes previamente pesadas, deje agitarse durante 7 minutos y agregue el siguiente reactivo" 
                },
            "15-201":{
                "descripcion": "Liquido incoloro; viscosidad Baja; libre de particulas",
                "peligro":"m",
                "indicacion":"Pese el 15-201 y agreguelo bajo agitación",
                "revision":"Agite durante 8-10 minutos"
                },
            "15-532":{
                "descripcion": "Polvo fino gris",
                "peligro":"0",
                "indicacion":"Pese el 15-532 y agreguelo bajo agitación",
                "revision":"Agite durante 5-7 minutos"
                },
            "15-564":{
                "descripcion": "Polvo Fino Blanco",
                "peligro":"",
                "indicacion":"Pese el 15-633 y agreguelo bajo agitación",
                "revision":"Agite durante 5-7 minutos"
                },
            "15-520":{
                "descripcion": "Polvo Fino Blanco",
                "peligro":"t",
                "indicacion":"Pese el 15-633 y agreguelo bajo agitación",
                "revision":"Agite durante 5-7 minutos"
                },
            "15-515":{
                "descripcion": "Polvo Fino Blanco",
                "peligro":"0",
                "indicacion":"Pese el 15-515 y agreguelo bajo agitación",
                "revision":"Agite durante 5-7 minutos"
                },
            "15-402":{
                "descripcion": "Liquido incoloro; viscosidad Baja; libre de particulas",
                "peligro":"t",
                "indicacion":"Pese el 15-417 y agregue bajo agitación. Deje agitar durante 10 minutos",
                "revision":"verifique que el liquido este libre de particulas"
                },
            "20-500":{
                "descripcion": "Pasta translucida",
                "peligro":"0",
                "indicacion":"Pese el 15-633 y agreguelo una vez la mezcla se haya vuelto muy viscosa o pesada",
                "revision":"verifique que el liquido este libre de particulas"
                },
            "Banner 3":"b8",
            "Equipo a utilizar": {
                "Equipo a utilizar":"Tina de Acero Inox, Dispersor o Tambo de proceso" 
                },
            "Envase":{
                "Envase_producto":"Envase indicado en la Orden de Producción" ,
                "Envase_control":"Porron de plastico"
                }
            }
        }
    def transferir_producto(self,nombre):
        pass        
##wb = Workbook()
##ws = wb.active
##ws.title = "Productos"
##headings = ['producto']+ list(productos["HB12"].keys())
##ws.append(headings)
clase_productos = Productos()
lista_productos = clase_productos.productos
##for producto in lista_productos.keys():# para ver las KEY's 
##    print(producto)
##    for reactivo in lista_productos[str(producto)].keys():
####        print(reactivo)
##        print(lista_productos[str(producto)][reactivo].values())
descripcion = lista_productos['HB12']['15-340-1']['descripcion']
print(descripcion)

# INFORMACION DE OS BANNERS

#   'b1', # BANNER DE PREPARACION 
#   'b2', # BANNER DE EPP 
#   'b3', # BANNER DE "PESE, CALCULE, ANOTE, AÑADA, TIRE/GUARDE" 
#   'b4', # BANNER DE "MEZCLAR EN 4 PARTES" 15-515
#   'b5', # BANNER DE advert. de explosion 
#   'b6', # ADVERTENCIA DEL TDI 
#   'b7', # CONSIDERACIONES BÁSICAS ANTES DE EMPEZAR EL PROCESO 
#   'b8', # MANTENGA LIMPIO SU REACTOR  
#   'b9', # PASE MUESTRA A CALIDAD
#   'b10', # Solicitar Fineza
#   'b11', # PASE MUESTRA A CALIDAD bote sucio
#   'b12', # EVITE DEJAR FUNCIONANDO MUCHO TIEMPO EL DISPERSOR
